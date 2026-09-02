"""Read-only checks before comparing the final FinanceBench dev review tables.

Place this file in noisy-rag-eval/src and run:
    python src/check_final_reviews.py

Dependency: openpyxl (reading XLSX only). No API calls; no file writes.
This validates structure, provenance and label consistency, not label truth.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from collections import Counter
from decimal import Decimal, InvalidOperation
from itertools import islice
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
EXPECTED_ROWS = 30
COMMON = {
    "sample_id", "question_type", "company", "question", "gold_answer",
    "predicted_answer", "confidence", "abstain", "reason",
    "manual_correct", "source_hallucination", "error_type",
}
ALIASES = {
    "predicted_answer": "new_predicted_answer",
    "confidence": "new_confidence",
    "abstain": "new_abstain",
    "reason": "new_reason",
}
ERROR_TYPES = {
    "correct", "wrong_numeric", "wrong_conclusion", "incomplete",
    "abstention", "dataset_issue",
}
FAILURE_STAGES = {
    "none", "retrieval", "generation", "retrieval_and_generation",
    "dataset", "unclear",
}


def empty(value):
    return value is None or (isinstance(value, str) and not value.strip())


def text(value):
    return "" if empty(value) else " ".join(str(value).split())


def equivalent(a, b):
    if text(a) == text(b):
        return True
    try:
        left, right = Decimal(text(a)), Decimal(text(b))
        return left.is_finite() and right.is_finite() and left == right
    except InvalidOperation:
        return False


def numeric_gold(value):
    """Parse a whole numeric gold cell without discarding currency/percent units.

    Accept only plain numbers, dollar amounts, or percentages. Commas must be
    valid three-digit thousands groups. Do not extract numbers from prose,
    convert units, round values, or treat a percentage as an unmarked number.
    """
    match = re.fullmatch(
        r"(?P<currency>\$?)\s*"
        r"(?P<number>[+-]?(?:[0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)(?:\.[0-9]+)?)"
        r"\s*(?P<percent>%?)",
        text(value),
    )
    if match is None or (match['currency'] and match['percent']):
        return None
    return (
        match['currency'],
        Decimal(match['number'].replace(',', '')),
        match['percent'],
    )


def gold_equivalent(a, b):
    if equivalent(a, b):
        return True
    left, right = numeric_gold(a), numeric_gold(b)
    return left is not None and right is not None and left == right


def binary(value, optional=False):
    if empty(value):
        if optional:
            return None
        raise ValueError("标签为空")
    token = text(value).lower()
    if optional and token in {"n/a", "na"}:
        return None
    if token == "true":
        return 1
    if token == "false":
        return 0
    try:
        number = Decimal(token)
        if number in (Decimal(0), Decimal(1)):
            return int(number)
    except InvalidOperation:
        pass
    raise ValueError(f"应为0或1（兼容TRUE/FALSE），实际为{value!r}")


def confidence(value):
    if isinstance(value, bool) or empty(value):
        raise ValueError(f"非法置信度：{value!r}")
    number = float(value)
    if not math.isfinite(number) or not 0 <= number <= 1:
        raise ValueError(f"置信度不在0~1内：{value!r}")
    return number


def read_workbook(path, role, selected_sheet=None):
    from openpyxl import load_workbook

    required = COMMON | (
        {"citation_support_correct", "failure_stage"} if role == "rag" else set()
    )
    if not path.is_file():
        raise ValueError(f"文件不存在：{path}")
    with path.open("rb") as stream:
        book = load_workbook(stream, read_only=True, data_only=True)
        try:
            if selected_sheet and selected_sheet not in book.sheetnames:
                raise ValueError(f"找不到工作表{selected_sheet!r}；已有：{book.sheetnames}")
            candidates = []
            header_hints = []
            for sheet in book.worksheets:
                if selected_sheet and sheet.title != selected_sheet:
                    continue
                for row_num, row in enumerate(islice(sheet.iter_rows(values_only=True), 20), 1):
                    headers = [text(value).lstrip("\ufeff") for value in row]
                    if "sample_id" not in headers:
                        continue
                    present = set(headers)
                    canonical = present | {
                        key for key, alias in ALIASES.items() if alias in present
                    }
                    header_hints.append((sheet.title, row_num, sorted(required - canonical)))
                    if not required - canonical:
                        candidates.append((sheet, row_num, headers))
            if len(candidates) != 1:
                raise ValueError(
                    f"{path.name} 应恰好识别到一个审核页，实际{len(candidates)}个。"
                    f"工作表：{book.sheetnames}；候选表头(页名,行号,缺失列)：{header_hints}。"
                    "若有多个审核页，可用 --llm-sheet 或 --rag-sheet 指定。"
                )
            sheet, header_row, headers = candidates[0]
            named = [header for header in headers if header]
            if len(named) != len(set(named)):
                raise ValueError(f"{path.name}/{sheet.title} 表头含重复列名")
            rows = []
            for row_num, values in enumerate(
                sheet.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True),
                header_row + 1,
            ):
                if all(empty(value) for value in values):
                    continue
                row = {key: value for key, value in zip(headers, values) if key}
                for key, alias in ALIASES.items():
                    if alias in row:
                        if key in row and not equivalent(row[key], row[alias]):
                            raise ValueError(f"{path.name} 第{row_num}行：{key}和{alias}不一致")
                        row[key] = row[alias]
                row["sample_id"] = text(row.get("sample_id"))
                row["_excel_row"] = row_num
                rows.append(row)
            print(f"读取：{path.name}\n  工作表={sheet.title!r}，表头=第{header_row}行，数据={len(rows)}条")
            return rows
        finally:
            book.close()


def read_jsonl(path):
    if not path.is_file():
        raise ValueError(f"文件不存在：{path}")
    rows = []
    with path.open(encoding="utf-8-sig") as stream:
        for number, line in enumerate(stream, 1):
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path.name} 第{number}行不是合法JSON") from exc
            if not isinstance(value, dict):
                raise ValueError(f"{path.name} 第{number}行不是JSON对象")
            rows.append(value)
    return rows


def index_rows(rows, label, errors):
    ids = [text(row.get("sample_id")) for row in rows]
    if len(rows) != EXPECTED_ROWS:
        errors.append(f"{label}：应为30条，实际{len(rows)}条")
    bad = [sid for sid in ids if not re.fullmatch(r"financebench_id_\d+", sid)]
    if bad:
        errors.append(f"{label}：缺失或非法样本ID：{bad}")
    duplicates = [sid for sid, count in Counter(ids).items() if count > 1]
    if duplicates:
        errors.append(f"{label}：重复ID：{duplicates}")
    return {text(row.get("sample_id")): row for row in rows}


def check_labels(rows, role, errors, warnings):
    for row in rows:
        where = f"{role}/{row['sample_id']}（Excel第{row['_excel_row']}行）"
        valid = True
        for field in ("manual_correct", "source_hallucination", "abstain"):
            try:
                row[field] = binary(row.get(field))
            except ValueError as exc:
                errors.append(f"{where} {field}：{exc}")
                valid = False
        try:
            row["confidence"] = confidence(row.get("confidence"))
        except (ValueError, TypeError) as exc:
            errors.append(f"{where} confidence：{exc}")
            valid = False
        if not valid:
            continue
        correct, abstain = row["manual_correct"], row["abstain"]
        kind = text(row.get("error_type"))
        if kind not in ERROR_TYPES:
            errors.append(f"{where} 非法error_type：{kind!r}")
        if (correct == 1) != (kind == "correct"):
            errors.append(f"{where} manual_correct与error_type=correct不一致")
        if abstain and (correct or kind not in {"abstention", "dataset_issue"}):
            errors.append(f"{where} 拒答应判0，error_type应为abstention（或有说明的dataset_issue）")
        if not abstain and kind == "abstention":
            errors.append(f"{where} 非拒答样本被标成abstention")
        if not abstain and empty(row.get("predicted_answer")):
            errors.append(f"{where} 非拒答样本的答案为空")
        if kind == "dataset_issue":
            warnings.append(f"{where} dataset_issue：后续需保留官方评分与清洁子集两个口径")
        if role != "rag":
            continue
        try:
            support = binary(row.get("citation_support_correct"), optional=True)
            row["citation_support_correct"] = support
        except ValueError as exc:
            errors.append(f"{where} citation_support_correct：{exc}")
            continue
        citation_key = next((key for key in ("model_citations", "citations") if key in row), None)
        has_citations = citation_key is not None and text(row[citation_key]).lower() not in {
            "", "[]", "none", "null", "nan",
        }
        if not abstain and support is None:
            errors.append(f"{where} 已作答，但citation_support_correct为空")
        if abstain and not has_citations and support is not None:
            errors.append(f"{where} 拒答且无引用，citation_support_correct应留空，不应填0或1")
        if abstain and has_citations:
            warnings.append(f"{where} 拒答中带有引用，请人工确认是否有待核验主张")
        stage = text(row.get("failure_stage"))
        if stage not in FAILURE_STAGES:
            errors.append(f"{where} 非法failure_stage：{stage!r}")
        if stage == "none" and (correct != 1 or support != 1 or row["source_hallucination"] != 0):
            errors.append(f"{where} failure_stage=none与答案/引用/来源幻觉标签矛盾")
        if stage == "unclear":
            warnings.append(f"{where} 失败阶段尚未确定，不应强行归类")


def compare_fields(left, right, fields, label, errors, notices=None):
    if set(left) != set(right):
        errors.append(
            f"{label} ID不一致：左侧缺{sorted(set(right)-set(left))}；"
            f"右侧缺{sorted(set(left)-set(right))}"
        )
    for sid in sorted(left.keys() & right.keys()):
        for field in fields:
            if field not in left[sid] or field not in right[sid]:
                errors.append(f"{label}/{sid} 缺少{field}")
                continue
            a, b = left[sid][field], right[sid][field]
            if equivalent(a, b):
                continue
            # Numeric display normalization is ONLY for gold metadata, never
            # for model answers/reasons or human labels. Preserve input cells.
            if field == "gold_answer" and gold_equivalent(a, b):
                if notices is not None:
                    notices.append(
                        f"{label}/{sid} gold_answer仅格式不同：{a!r} 与 {b!r}"
                        "（数值及货币/百分号标记一致；保留原值）"
                    )
                continue
            detail = f"：左侧{a!r}；右侧{b!r}" if field == "gold_answer" else ""
            errors.append(f"{label}/{sid} 的{field}不一致{detail}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=ROOT)
    parser.add_argument("--llm-review", default="results/metrics/llm_only_dev_qa_protocol_v1_manual_review.xlsx")
    parser.add_argument("--rag-review", default="results/metrics/dense_rag_dev_manual_review.xlsx")
    parser.add_argument("--llm-raw", default="results/raw_outputs/llm_only_dev_qa_protocol_v1.jsonl")
    parser.add_argument("--dev-data", default="data/processed/dev_30.jsonl")
    parser.add_argument("--llm-sheet")
    parser.add_argument("--rag-sheet")
    args = parser.parse_args()
    resolve = lambda name: (args.project_root / name).resolve()
    errors, warnings = [], []

    print("最终审核表检查（只读；不调用API；不生成或覆盖任何文件）")
    print("脚本版本：1.1（兼容gold数值格式差异；不改变数值或单位）")
    llm_rows = read_workbook(resolve(args.llm_review), "llm", args.llm_sheet)
    rag_rows = read_workbook(resolve(args.rag_review), "rag", args.rag_sheet)
    raw_rows = read_jsonl(resolve(args.llm_raw))
    dev_rows = read_jsonl(resolve(args.dev_data))
    llm = index_rows(llm_rows, "LLM审核表", errors)
    rag = index_rows(rag_rows, "RAG审核表", errors)
    raw = index_rows(raw_rows, "新版LLM原始结果", errors)
    dev = index_rows(dev_rows, "开发集", errors)

    check_labels(llm_rows, "llm", errors, warnings)
    check_labels(rag_rows, "rag", errors, warnings)
    metadata = ("question", "company", "question_type", "gold_answer")
    compare_fields(llm, rag, metadata, "LLM与RAG表格对齐", errors, warnings)
    compare_fields(llm, raw, metadata + ("predicted_answer", "reason"), "LLM表与新版原始结果", errors, warnings)
    compare_fields(raw, dev, metadata, "新版LLM原始结果与开发集", errors, warnings)

    expected = {"model": "deepseek-v4-flash", "prompt_version": "qa_protocol_v1", "mode": "llm_only", "max_tokens": 600}
    for row in raw_rows:
        sid = text(row.get("sample_id"))
        for field, value in expected.items():
            if row.get(field) != value:
                errors.append(f"原始结果/{sid} {field}不符合预期{value!r}")
        if row.get("citations") != []:
            errors.append(f"原始结果/{sid} LLM Only citations不是空列表")
        if "answer" in row and not equivalent(row["answer"], row.get("predicted_answer")):
            errors.append(f"原始结果/{sid} answer和predicted_answer不一致")
        try:
            raw_abstain, raw_confidence = binary(row.get("abstain")), confidence(row.get("confidence"))
            if sid in llm:
                if binary(llm[sid].get("abstain")) != raw_abstain:
                    errors.append(f"LLM表/{sid} abstain与新版原始结果不一致")
                if not math.isclose(confidence(llm[sid].get("confidence")), raw_confidence, abs_tol=1e-9):
                    errors.append(f"LLM表/{sid} confidence与新版原始结果不一致")
        except (TypeError, ValueError) as exc:
            errors.append(f"原始结果/表格/{sid}：{exc}")
    hashes = {text(row.get("dev_data_sha256")) for row in raw_rows}
    if len(hashes) != 1 or "" in hashes:
        errors.append("新版LLM原始结果混用了开发集哈希，或缺少哈希")
    else:
        current_hash = hashlib.sha256(resolve(args.dev_data).read_bytes()).hexdigest()
        if hashes != {current_hash}:
            warnings.append("开发集字节哈希与运行时不同（可能是换行/格式改变）；问题、公司、类型和gold已另做逐题对齐检查")

    for label, rows in (("LLM", llm_rows), ("RAG", rag_rows)):
        balance = Counter(text(row.get("question_type")) for row in rows)
        expected_balance = {"novel-generated": 10, "domain-relevant": 10, "metrics-generated": 10}
        if dict(balance) != expected_balance:
            errors.append(f"{label}问题类型数量异常：{dict(balance)}")
    print("\n提示：")
    for warning in warnings:
        print("- " + warning)
    print("- 空白notes不会被视为漏标；RAG拒答的引用支持空白不会被转成0。")
    print("- 只检查标签的一致性，不替代逐题人工审核。")
    print("- 尚未核验RAG原始请求参数；仅凭审核表不能证明全部运行设置相同。")
    if errors:
        print(f"\n检查未通过：{len(errors)}项问题（没有修改任何文件）")
        for error in errors:
            print("- " + error)
        return 1

    print("\n检查通过：两表均30条、ID一一对应、必填标签完整且逻辑一致。")
    print("LLM审核表的答案、解释、置信度、拒答状态与qa_protocol_v1原始结果一致。")
    for label, rows in (("LLM Only", llm_rows), ("Dense RAG", rag_rows)):
        print(f"\n{label}：")
        print(f"  manual_correct：{dict(sorted(Counter(row['manual_correct'] for row in rows).items()))}")
        print(f"  source_hallucination：{dict(sorted(Counter(row['source_hallucination'] for row in rows).items()))}")
        print(f"  abstain：{dict(sorted(Counter(row['abstain'] for row in rows).items()))}")
        if label == "Dense RAG":
            null_count = sum(row["citation_support_correct"] is None for row in rows)
            print(f"  citation_support_correct空白：{null_count}")
    print("\n下一步：根据本次检查结果导出规范CSV，再计算正式对比指标。")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except ImportError as exc:
        print(f"缺少Excel读取依赖：{exc}\n请先执行：python -m pip install openpyxl")
        raise SystemExit(2)
    except (OSError, ValueError, KeyError) as exc:
        print(f"检查已停止：{exc}\n没有修改任何文件，请把完整输出发回。")
        raise SystemExit(2)
